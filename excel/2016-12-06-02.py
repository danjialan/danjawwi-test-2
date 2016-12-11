import openpyxl
import openpyxl.cell

wb = openpyxl.load_workbook('C:/Users/danjawwi/Desktop/python/excel/2.xlsx')

ws = wb.active
print(ws.title)
a = ws.cell(row = 1,column = 2)
print(a.value)

print(ws.max_row)
print(ws.max_column)

for i in range(1,5):
    print(i , ws.cell(row = i , column = 2).value)

ws2 = wb.get_sheet_by_name('test1')
list(ws2['A1':'B5'])
for a in ws2['A1':'B5']:
    for b in a :
        print(b.coordinate,b.value)
    print('---end of row---')