from openpyxl import Workbook
from openpyxl import load_workbook
from openpyxl.styles import PatternFill, Border, Side, Alignment, Protection, Font, Color, Fill
from openpyxl.styles import colors
from openpyxl.styles import Fill,fills
from openpyxl.formatting.rule import ColorScaleRule

wb = load_workbook('C:/Users/danjawwi/Desktop/python/excel/1.xlsx')
ws = wb['Sheet1']
sheet_names = wb.get_sheet_names()
print(type(sheet_names))
print(sheet_names)
ws2 = wb.get_sheet_by_name(sheet_names[1])
ws3 = wb.create_sheet(title='test1',index = 0)
ws4 = wb.create_sheet(title='test2')
wb.create_sheet(title='test3')
ws3['A4'] = 5
a = ws2['A4']
b = ws2.cell(row = 4, column = 1)
print(str(a))
print(str(b))
wb.save(filename= 'C:/Users/danjawwi/Desktop/python/excel/2.xlsx')